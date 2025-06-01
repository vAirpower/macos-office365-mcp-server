"""Excel Controller for Office 365 MCP Server
Handles all Excel automation operations on macOS.
"""

import asyncio
import uuid
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Union, Tuple
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import coordinate_from_string

import sys
import os
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from integrations.applescript_bridge import AppleScriptBridge
from utils.logger import setup_logger

logger = setup_logger(__name__)

class ExcelController:
    """Controller for Excel operations using openpyxl."""
    
    def __init__(self):
        self.applescript = AppleScriptBridge()
        self.active_workbooks: Dict[str, Dict[str, Any]] = {}
        # Use system temp directory with a subdirectory
        self.temp_dir = Path(tempfile.gettempdir()) / "office365_mcp"
        self.temp_dir.mkdir(parents=True, exist_ok=True)
    
    async def create_workbook(
        self,
        title: str = "New Workbook",
        template_path: Optional[str] = None
    ) -> Dict[str, Any]:
        """Create a new Excel workbook.
        
        Args:
            title: Workbook title
            template_path: Optional template file path
            
        Returns:
            Dict with workbook metadata
        """
        try:
            workbook_id = str(uuid.uuid4())
            
            # Create workbook using openpyxl
            if template_path and Path(template_path).exists():
                wb = load_workbook(template_path)
            else:
                wb = Workbook()
            
            # Set title in first cell if new workbook
            if not template_path:
                ws = wb.active
                ws.title = "Sheet1"
                ws['A1'] = title
                ws['A1'].font = Font(bold=True, size=14)
            
            # Save temporary file
            temp_file = self.temp_dir / f"{workbook_id}.xlsx"
            wb.save(str(temp_file))
            
            # Try to open in Excel via AppleScript
            applescript_success = False
            try:
                # Note: Would need to implement open_excel_file in AppleScriptBridge
                # await self.applescript.open_excel_file(str(temp_file))
                applescript_success = False  # For now
            except Exception as e:
                logger.warning(f"Could not open in Excel app: {e}")
            
            # Store workbook metadata
            workbook_data = {
                "workbook_id": workbook_id,
                "title": title,
                "file_path": str(temp_file),
                "worksheet_count": len(wb.worksheets),
                "active_sheet": wb.active.title,
                "applescript_available": applescript_success,
                "created_at": asyncio.get_event_loop().time()
            }
            
            self.active_workbooks[workbook_id] = {
                "metadata": workbook_data,
                "workbook_object": wb,
                "worksheets": {ws.title: ws for ws in wb.worksheets}
            }
            
            logger.info(f"Created workbook: {title} ({workbook_id})")
            return workbook_data
            
        except Exception as e:
            logger.error(f"Failed to create workbook: {e}")
            raise
    
    async def add_worksheet(
        self,
        workbook_id: str,
        sheet_name: str,
        position: Optional[int] = None
    ) -> Dict[str, Any]:
        """Add a new worksheet to a workbook.
        
        Args:
            workbook_id: ID of the workbook
            sheet_name: Name for the new sheet
            position: Position to insert sheet
            
        Returns:
            Dict with worksheet metadata
        """
        try:
            if workbook_id not in self.active_workbooks:
                raise ValueError(f"Workbook {workbook_id} not found")
            
            wb = self.active_workbooks[workbook_id]["workbook_object"]
            
            # Create worksheet
            if position is not None:
                ws = wb.create_sheet(sheet_name, position)
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Update worksheets dict
            self.active_workbooks[workbook_id]["worksheets"][sheet_name] = ws
            
            # Update metadata
            self.active_workbooks[workbook_id]["metadata"]["worksheet_count"] = len(wb.worksheets)
            
            # Save workbook
            temp_file = self.active_workbooks[workbook_id]["metadata"]["file_path"]
            wb.save(temp_file)
            
            logger.info(f"Added worksheet '{sheet_name}' to workbook {workbook_id}")
            return {
                "status": "success",
                "workbook_id": workbook_id,
                "sheet_name": sheet_name,
                "sheet_index": wb.worksheets.index(ws)
            }
            
        except Exception as e:
            logger.error(f"Failed to add worksheet: {e}")
            raise
    
    async def write_cell(
        self,
        workbook_id: str,
        sheet_name: str,
        cell: str,
        value: Any,
        formatting: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Write data to a specific cell.
        
        Args:
            workbook_id: ID of the workbook
            sheet_name: Name of the worksheet
            cell: Cell reference (e.g., 'A1')
            value: Value to write
            formatting: Optional formatting options
            
        Returns:
            Dict with operation status
        """
        try:
            if workbook_id not in self.active_workbooks:
                raise ValueError(f"Workbook {workbook_id} not found")
            
            ws = self.active_workbooks[workbook_id]["worksheets"].get(sheet_name)
            if not ws:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            # Write value
            ws[cell] = value
            
            # Apply formatting if provided
            if formatting:
                await self._apply_cell_formatting(ws[cell], formatting)
            
            # Save workbook
            wb = self.active_workbooks[workbook_id]["workbook_object"]
            temp_file = self.active_workbooks[workbook_id]["metadata"]["file_path"]
            wb.save(temp_file)
            
            logger.info(f"Wrote value to cell {cell} in sheet '{sheet_name}'")
            return {
                "status": "success",
                "workbook_id": workbook_id,
                "sheet_name": sheet_name,
                "cell": cell,
                "value": str(value)
            }
            
        except Exception as e:
            logger.error(f"Failed to write cell: {e}")
            raise
    
    async def write_range(
        self,
        workbook_id: str,
        sheet_name: str,
        start_cell: str,
        data: List[List[Any]],
        formatting: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Write data to a range of cells.
        
        Args:
            workbook_id: ID of the workbook
            sheet_name: Name of the worksheet
            start_cell: Starting cell reference
            data: 2D list of values
            formatting: Optional formatting options
            
        Returns:
            Dict with operation status
        """
        try:
            if workbook_id not in self.active_workbooks:
                raise ValueError(f"Workbook {workbook_id} not found")
            
            ws = self.active_workbooks[workbook_id]["worksheets"].get(sheet_name)
            if not ws:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            # Parse start cell
            col_letter, start_row = coordinate_from_string(start_cell)
            start_col = column_index_from_string(col_letter)
            
            # Write data
            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    cell = ws.cell(row=start_row + row_idx, column=start_col + col_idx)
                    cell.value = value
                    
                    # Apply formatting if provided
                    if formatting:
                        await self._apply_cell_formatting(cell, formatting)
            
            # Save workbook
            wb = self.active_workbooks[workbook_id]["workbook_object"]
            temp_file = self.active_workbooks[workbook_id]["metadata"]["file_path"]
            wb.save(temp_file)
            
            end_row = start_row + len(data) - 1
            end_col = start_col + max(len(row) for row in data) - 1
            end_cell = f"{get_column_letter(end_col)}{end_row}"
            
            logger.info(f"Wrote data to range {start_cell}:{end_cell}")
            return {
                "status": "success",
                "workbook_id": workbook_id,
                "sheet_name": sheet_name,
                "range": f"{start_cell}:{end_cell}",
                "rows_written": len(data)
            }
            
        except Exception as e:
            logger.error(f"Failed to write range: {e}")
            raise
    
    async def add_formula(
        self,
        workbook_id: str,
        sheet_name: str,
        cell: str,
        formula: str
    ) -> Dict[str, Any]:
        """Add a formula to a cell.
        
        Args:
            workbook_id: ID of the workbook
            sheet_name: Name of the worksheet
            cell: Cell reference
            formula: Excel formula (e.g., '=SUM(A1:A10)')
            
        Returns:
            Dict with operation status
        """
        try:
            if workbook_id not in self.active_workbooks:
                raise ValueError(f"Workbook {workbook_id} not found")
            
            ws = self.active_workbooks[workbook_id]["worksheets"].get(sheet_name)
            if not ws:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            # Ensure formula starts with =
            if not formula.startswith('='):
                formula = '=' + formula
            
            # Write formula
            ws[cell] = formula
            
            # Save workbook
            wb = self.active_workbooks[workbook_id]["workbook_object"]
            temp_file = self.active_workbooks[workbook_id]["metadata"]["file_path"]
            wb.save(temp_file)
            
            logger.info(f"Added formula to cell {cell}")
            return {
                "status": "success",
                "workbook_id": workbook_id,
                "sheet_name": sheet_name,
                "cell": cell,
                "formula": formula
            }
            
        except Exception as e:
            logger.error(f"Failed to add formula: {e}")
            raise
    
    async def create_chart(
        self,
        workbook_id: str,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        chart_title: str = "",
        position: str = "E5"
    ) -> Dict[str, Any]:
        """Create a chart from data.
        
        Args:
            workbook_id: ID of the workbook
            sheet_name: Name of the worksheet
            chart_type: Type of chart (bar, line, pie)
            data_range: Data range (e.g., 'A1:B10')
            chart_title: Title for the chart
            position: Cell position for chart
            
        Returns:
            Dict with operation status
        """
        try:
            if workbook_id not in self.active_workbooks:
                raise ValueError(f"Workbook {workbook_id} not found")
            
            ws = self.active_workbooks[workbook_id]["worksheets"].get(sheet_name)
            if not ws:
                raise ValueError(f"Worksheet '{sheet_name}' not found")
            
            # Parse data range
            from openpyxl.utils import range_boundaries
            min_col, min_row, max_col, max_row = range_boundaries(data_range)
            
            # Create appropriate chart
            if chart_type.lower() == "bar":
                chart = BarChart()
            elif chart_type.lower() == "line":
                chart = LineChart()
            elif chart_type.lower() == "pie":
                chart = PieChart()
            else:
                raise ValueError(f"Unsupported chart type: {chart_type}")
            
            # Set chart title
            if chart_title:
                chart.title = chart_title
            
            # Add data
            data = Reference(ws, min_col=min_col, min_row=min_row, 
                           max_col=max_col, max_row=max_row)
            chart.add_data(data, titles_from_data=True)
            
            # Add chart to worksheet
            ws.add_chart(chart, position)
            
            # Save workbook
            wb = self.active_workbooks[workbook_id]["workbook_object"]
            temp_file = self.active_workbooks[workbook_id]["metadata"]["file_path"]
            wb.save(temp_file)
            
            logger.info(f"Created {chart_type} chart at {position}")
            return {
                "status": "success",
                "workbook_id": workbook_id,
                "sheet_name": sheet_name,
                "chart_type": chart_type,
                "position": position,
                "data_range": data_range
            }
            
        except Exception as e:
            logger.error(f"Failed to create chart: {e}")
            raise
    
    async def save_workbook(
        self,
        workbook_id: str,
        file_path: str,
        format: str = "xlsx"
    ) -> Dict[str, Any]:
        """Save a workbook to file.
        
        Args:
            workbook_id: ID of the workbook
            file_path: Path to save the file
            format: File format (xlsx, xlsm, etc.)
            
        Returns:
            Dict with operation status
        """
        try:
            if workbook_id not in self.active_workbooks:
                raise ValueError(f"Workbook {workbook_id} not found")
            
            wb = self.active_workbooks[workbook_id]["workbook_object"]
            save_path = Path(file_path)
            
            # Ensure directory exists
            save_path.parent.mkdir(parents=True, exist_ok=True)
            
            # Save with appropriate extension
            if not save_path.suffix:
                save_path = save_path.with_suffix(f".{format}")
            
            wb.save(str(save_path))
            
            # Update metadata
            self.active_workbooks[workbook_id]["metadata"]["file_path"] = str(save_path)
            
            logger.info(f"Saved workbook to {save_path}")
            return {
                "status": "success",
                "workbook_id": workbook_id,
                "file_path": str(save_path),
                "format": format
            }
            
        except Exception as e:
            logger.error(f"Failed to save workbook: {e}")
            raise
    
    async def _apply_cell_formatting(
        self,
        cell,
        formatting: Dict[str, Any]
    ) -> None:
        """Apply formatting to a cell.
        
        Args:
            cell: Cell object
            formatting: Formatting options
        """
        try:
            # Font formatting
            font_kwargs = {}
            if "bold" in formatting:
                font_kwargs["bold"] = formatting["bold"]
            if "italic" in formatting:
                font_kwargs["italic"] = formatting["italic"]
            if "font_size" in formatting:
                font_kwargs["size"] = formatting["font_size"]
            if "font_color" in formatting:
                font_kwargs["color"] = formatting["font_color"]
            if "font_name" in formatting:
                font_kwargs["name"] = formatting["font_name"]
            
            if font_kwargs:
                cell.font = Font(**font_kwargs)
            
            # Fill (background color)
            if "bg_color" in formatting:
                cell.fill = PatternFill(start_color=formatting["bg_color"],
                                      end_color=formatting["bg_color"],
                                      fill_type="solid")
            
            # Alignment
            align_kwargs = {}
            if "horizontal" in formatting:
                align_kwargs["horizontal"] = formatting["horizontal"]
            if "vertical" in formatting:
                align_kwargs["vertical"] = formatting["vertical"]
            if "wrap_text" in formatting:
                align_kwargs["wrap_text"] = formatting["wrap_text"]
            
            if align_kwargs:
                cell.alignment = Alignment(**align_kwargs)
            
            # Border
            if "border" in formatting and formatting["border"]:
                side = Side(style='thin')
                cell.border = Border(left=side, right=side, top=side, bottom=side)
            
        except Exception as e:
            logger.warning(f"Failed to apply some cell formatting: {e}")
    
    async def get_workbook_info(self, workbook_id: str) -> Dict[str, Any]:
        """Get information about a workbook.
        
        Args:
            workbook_id: ID of the workbook
            
        Returns:
            Dict with workbook information
        """
        if workbook_id not in self.active_workbooks:
            raise ValueError(f"Workbook {workbook_id} not found")
        
        return self.active_workbooks[workbook_id]["metadata"]
    
    async def list_workbooks(self) -> List[Dict[str, Any]]:
        """List all active workbooks.
        
        Returns:
            List of workbook metadata
        """
        return [data["metadata"] for data in self.active_workbooks.values()]
    
    async def list_worksheets(self, workbook_id: str) -> List[str]:
        """List all worksheets in a workbook.
        
        Args:
            workbook_id: ID of the workbook
            
        Returns:
            List of worksheet names
        """
        if workbook_id not in self.active_workbooks:
            raise ValueError(f"Workbook {workbook_id} not found")
        
        wb = self.active_workbooks[workbook_id]["workbook_object"]
        return [ws.title for ws in wb.worksheets]